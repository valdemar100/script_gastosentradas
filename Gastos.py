import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

ARQUIVO = "gastos.xlsx"

class App:
    def __init__(self, root):
        self.root = root
        self.root.title("Controle de Gastos Diários")

        # Inicializa o arquivo Excel se não existir
        if not os.path.exists(ARQUIVO):
            wb = Workbook()
            ws = wb.active
            ws.append(["Data", "Descrição", "Tipo", "Valor"])
            wb.save(ARQUIVO)

        # Cria os widgets
        self.criar_widgets()

    def criar_widgets(self):
        tk.Label(self.root, text="Data (DD/MM/AAAA):").grid(row=0, column=0, sticky="e")
        self.data_entry = tk.Entry(self.root)
        self.data_entry.grid(row=0, column=1)

        tk.Label(self.root, text="Descrição:").grid(row=1, column=0, sticky="e")
        self.descricao_entry = tk.Entry(self.root)
        self.descricao_entry.grid(row=1, column=1)

        tk.Label(self.root, text="Tipo:").grid(row=2, column=0, sticky="e")
        self.tipo_var = tk.StringVar()
        tipo_menu = ttk.Combobox(self.root, textvariable=self.tipo_var, values=["Entrada", "Gasto"])
        tipo_menu.grid(row=2, column=1)

        tk.Label(self.root, text="Valor:").grid(row=3, column=0, sticky="e")
        self.valor_entry = tk.Entry(self.root)
        self.valor_entry.grid(row=3, column=1)
        
        enviar_btn = tk.Button(self.root, text="Enviar", command=self.salvar_dado)
        enviar_btn.grid(row=4, column=0, columnspan=2, pady=10)
        
        totais_btn = tk.Button(self.root, text="Ver saldo total", command=self.ver_totais)
        totais_btn.grid(row=5, column=0, columnspan=2, pady=5)
        
    # Salva os dados digitados no arquivo excel
    def salvar_dado(self):
        data = self.data_entry.get()
        descricao = self.descricao_entry.get()
        valor = self.valor_entry.get()
        tipo = self.tipo_var.get()

        if not data or not descricao or not valor or not tipo:
            messagebox.showerror("Erro", "Preencha todos os campos.")
            return

        try:
            float(valor)
            datetime.strptime(data, "%d/%m/%Y")
        except:
            messagebox.showerror("Erro", "Data deve estar no formato DD/MM/AAAA e valor numérico.")
            return

        wb = load_workbook(ARQUIVO)
        ws = wb.active
        ws.append([data, descricao, tipo, float(valor)])
        wb.save(ARQUIVO)

        messagebox.showinfo("Sucesso", "Registro salvo com sucesso!")

        self.data_entry.delete(0, tk.END)
        self.descricao_entry.delete(0, tk.END)
        self.valor_entry.delete(0, tk.END)
    
    def ver_totais(self):
        wb = load_workbook(ARQUIVO)
        ws = wb.active

        total_gastos = 0
        total_entradas = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            tipo = row[2]
            valor = row[3]
            if tipo == "Gasto":
                total_gastos += valor
            elif tipo == "Entrada":
                total_entradas += valor

        messagebox.showinfo("Totais", f"Total de Entradas: R$ {total_entradas:.2f}\nTotal de Gastos: R$ {total_gastos:.2f}")

# Executa o programa
if __name__ == "__main__":
    root = tk.Tk()
    app = App(root)
    root.mainloop()
